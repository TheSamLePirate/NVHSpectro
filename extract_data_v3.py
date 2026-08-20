import cv2
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

def get_plot_area_exact(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=100, minLineLength=100, maxLineGap=10)
    
    h, w = image.shape[:2]
    min_x, max_x = w, 0
    min_y, max_y = h, 0
    
    if lines is not None:
        for line in lines:
            x1, y1, x2, y2 = line.flatten()
            if abs(y1 - y2) < 5: # horizontal
                if x1 < min_x: min_x = x1
                if x2 > max_x: max_x = x2
                if y1 > max_y: max_y = y1
                if y1 < min_y: min_y = y1
            if abs(x1 - x2) < 5: # vertical
                if y1 < min_y: min_y = y1
                if y2 > max_y: max_y = y2
                if x1 > max_x: max_x = x1
                if x1 < min_x: min_x = x1
                
    # fallback to almost full image if lines not found
    if min_x >= max_x or min_y >= max_y:
        return 5, 5, w-10, h-10
        
    return min_x, min_y, max_x - min_x, max_y - min_y

def extract_red_curve_v3(img_path, target_rpms, rpm_min, rpm_max, db_min=120, db_max=138):
    img = cv2.imread(img_path)
    if img is None: return [None]*len(target_rpms)
    
    # 1. Upscale for precision
    scale = 4
    h, w = img.shape[:2]
    img_large = cv2.resize(img, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)
    
    # 2. Get exact plot area (scaled)
    x_plot, y_plot, w_plot, h_plot = get_plot_area_exact(img)
    x_plot *= scale
    y_plot *= scale
    w_plot *= scale
    h_plot *= scale
    
    # 3. Color extraction (Red)
    hsv = cv2.cvtColor(img_large, cv2.COLOR_BGR2HSV)
    mask1 = cv2.inRange(hsv, np.array([0, 70, 50]), np.array([10, 255, 255]))
    mask2 = cv2.inRange(hsv, np.array([170, 70, 50]), np.array([180, 255, 255]))
    mask = mask1 | mask2
    
    points = np.where(mask > 0)
    if len(points[0]) == 0:
        return [None]*len(target_rpms)
        
    pts_x = points[1]
    pts_y = points[0]
    
    # 4. Map to physical values
    # x_plot is rpm_min, x_plot + w_plot is rpm_max
    rpms = rpm_min + (pts_x - x_plot) / w_plot * (rpm_max - rpm_min)
    # y_plot (top) is db_max, y_plot + h_plot (bottom) is db_min
    dbs = db_max - (pts_y - y_plot) / h_plot * (db_max - db_min)
    
    # 5. Get upper limit (max dB) for each RPM
    max_db_at_rpm = {}
    for r, d in zip(rpms, dbs):
        # bin to 1 RPM for smoothing
        r_bin = round(r)
        if r_bin not in max_db_at_rpm or d > max_db_at_rpm[r_bin]:
            max_db_at_rpm[r_bin] = d
            
    if not max_db_at_rpm: return [None]*len(target_rpms)
    
    r_keys = sorted(list(max_db_at_rpm.keys()))
    d_vals = [max_db_at_rpm[k] for k in r_keys]
    
    # 6. Interpolate exactly at target RPMs
    interpolated_dbs = np.interp(target_rpms, r_keys, d_vals, left=np.nan, right=np.nan)
    return [round(val, 1) if not np.isnan(val) else None for val in interpolated_dbs]

def main():
    slides = {
        4: ('BFC1', 'rDL', 1000, 4500),
        5: ('BFC1', 'rDH', 4500, 13500),
        6: ('BFC1', 'rCL', 1000, 4500),
        7: ('BFC1', 'rCH', 4500, 13500),
        9: ('BFC2', 'rDL', 1000, 4500),
        10: ('BFC2', 'rDH', 4500, 13500),
        11: ('BFC2', 'rCL', 1000, 4500),
        12: ('BFC2', 'rCH', 4500, 13500)
    }
    
    harmonics = ['H6', 'H7.4', 'H14.8', 'H18', 'H28', 'H32', 'H54', 'H56']
    
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Donnees_Brutes"
    
    # All target RPMs across all cycles
    all_rpms = set()
    for _, _, rmin, rmax in slides.values():
        all_rpms.update(range(rmin, rmax + 250, 250))
    sorted_rpms = sorted(list(all_rpms))
    
    headers = ['BFC', 'Cycle', 'Harmonique'] + [f"{rpm} RPM" for rpm in sorted_rpms]
    ws_data.append(headers)
    
    all_data = []
    img_dir = "extracted_images"
    
    for slide_num, (bfc, cycle, rmin, rmax) in slides.items():
        # Specific target RPMs for this cycle
        cycle_rpms = list(range(rmin, rmax + 250, 250))
        
        for i, harm in enumerate(harmonics):
            img_name = f"slide_{slide_num}_img_{i+1}.png"
            img_path = os.path.join(img_dir, img_name)
            
            if os.path.exists(img_path):
                # Assuming 120-138 dB for now since OCR failed.
                dbs = extract_red_curve_v3(img_path, cycle_rpms, rmin, rmax, 120, 138)
            else:
                dbs = [None]*len(cycle_rpms)
                
            db_map = dict(zip(cycle_rpms, dbs))
            
            row = [bfc, cycle, harm]
            for rpm in sorted_rpms:
                row.append(db_map.get(rpm, None))
            ws_data.append(row)
            
            all_data.append({
                'bfc': bfc,
                'cycle': cycle,
                'harm': harm,
                'rpms': cycle_rpms,
                'dbs': dbs
            })
            
    ws_charts = wb.create_sheet(title="Graphiques")
    
    chart_col = 1
    chart_row = 1
    
    for cycle in ['rDL', 'rDH', 'rCL', 'rCH']:
        for harm in harmonics:
            d1 = next((d for d in all_data if d['cycle'] == cycle and d['harm'] == harm and d['bfc'] == 'BFC1'), None)
            d2 = next((d for d in all_data if d['cycle'] == cycle and d['harm'] == harm and d['bfc'] == 'BFC2'), None)
            
            if not d1 and not d2: continue
            
            chart = ScatterChart()
            chart.title = f"{cycle} - {harm}"
            chart.style = 13
            chart.x_axis.title = 'RPM'
            chart.y_axis.title = 'dB'
            
            idx1 = all_data.index(d1) + 2 if d1 else -1
            idx2 = all_data.index(d2) + 2 if d2 else -1
            
            xvalues = Reference(ws_data, min_col=4, min_row=1, max_col=4+len(sorted_rpms)-1)
            
            if idx1 != -1 and any(v is not None for v in d1['dbs']):
                values = Reference(ws_data, min_col=4, min_row=idx1, max_col=4+len(sorted_rpms)-1)
                series = Series(values, xvalues, title_from_data=False, title='BFC1')
                chart.series.append(series)
                
            if idx2 != -1 and any(v is not None for v in d2['dbs']):
                values = Reference(ws_data, min_col=4, min_row=idx2, max_col=4+len(sorted_rpms)-1)
                series = Series(values, xvalues, title_from_data=False, title='BFC2')
                chart.series.append(series)
                
            cell = ws_charts.cell(row=chart_row, column=chart_col)
            ws_charts.add_chart(chart, cell.coordinate)
            
            chart_col += 10
            if chart_col > 30:
                chart_col = 1
                chart_row += 15

    output_path = r"C:\Users\Louis\Downloads\Export_NVH_Courbes_V3.xlsx"
    wb.save(output_path)
    print(f"Saved to {output_path}")

if __name__ == '__main__':
    main()
