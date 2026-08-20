import cv2
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

def get_plot_area(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return 0, 0, image.shape[1], image.shape[0]
    
    # find largest bounding rect
    max_area = 0
    best_rect = (0, 0, image.shape[1], image.shape[0])
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        area = w * h
        if area > max_area and area > 0.1 * image.shape[0] * image.shape[1]:
            max_area = area
            best_rect = (x, y, w, h)
            
    # if no large contour is found, just return the whole image minus some margin
    if max_area == 0:
        h, w = image.shape[:2]
        return int(w*0.1), int(h*0.1), int(w*0.8), int(h*0.8)
    return best_rect

def extract_red_curve(img_path, target_rpms):
    img = cv2.imread(img_path)
    if img is None:
        return [None]*len(target_rpms)
        
    x_plot, y_plot, w_plot, h_plot = get_plot_area(img)
    
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    
    lower_red2 = np.array([170, 70, 50])
    upper_red2 = np.array([180, 255, 255])
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    
    mask = mask1 | mask2
    
    points = np.where(mask > 0)
    if len(points[0]) == 0:
        return [None]*len(target_rpms)
        
    # points[1] is x (col), points[0] is y (row)
    # map x to RPM (1000 to 4000)
    rpms = 1000 + (points[1] - x_plot) / w_plot * (4000 - 1000)
    # map y to dB (138 to 120, since y=0 is top)
    dbs = 138 - (points[0] - y_plot) / h_plot * (138 - 120)
    
    # We want max dB for each RPM
    max_db_at_rpm = {}
    for r, d in zip(rpms, dbs):
        r_bin = round(r)
        if r_bin not in max_db_at_rpm or d > max_db_at_rpm[r_bin]:
            max_db_at_rpm[r_bin] = d
            
    if not max_db_at_rpm:
         return [None]*len(target_rpms)
         
    r_keys = sorted(list(max_db_at_rpm.keys()))
    d_vals = [max_db_at_rpm[k] for k in r_keys]
    
    # Interpolate at target RPMs
    interpolated_dbs = np.interp(target_rpms, r_keys, d_vals, left=np.nan, right=np.nan)
    
    return [round(val, 1) if not np.isnan(val) else None for val in interpolated_dbs]

def main():
    target_rpms = list(range(1000, 4250, 250))
    
    slides = {
        4: ('BFC1', 'rDL'),
        5: ('BFC1', 'rDH'),
        6: ('BFC1', 'rCL'),
        7: ('BFC1', 'rCH'),
        9: ('BFC2', 'rDL'),
        10: ('BFC2', 'rDH'),
        11: ('BFC2', 'rCL'),
        12: ('BFC2', 'rCH')
    }
    
    harmonics = ['H6', 'H7.4', 'H14.8', 'H18', 'H28', 'H32', 'H54', 'H56']
    
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Donnees_Brutes"
    
    headers = ['BFC', 'Cycle', 'Harmonique'] + [f"{rpm} RPM" for rpm in target_rpms]
    ws_data.append(headers)
    
    all_data = []
    
    img_dir = "extracted_images"
    
    for slide_num, (bfc, cycle) in slides.items():
        for i, harm in enumerate(harmonics):
            img_name = f"slide_{slide_num}_img_{i+1}.png"
            img_path = os.path.join(img_dir, img_name)
            
            if os.path.exists(img_path):
                dbs = extract_red_curve(img_path, target_rpms)
            else:
                dbs = [None]*len(target_rpms)
                
            row = [bfc, cycle, harm] + dbs
            ws_data.append(row)
            
            all_data.append({
                'bfc': bfc,
                'cycle': cycle,
                'harm': harm,
                'rpms': target_rpms,
                'dbs': dbs
            })
            
    ws_charts = wb.create_sheet(title="Graphiques")
    
    chart_col = 1
    chart_row = 1
    
    for cycle in ['rDL', 'rDH', 'rCL', 'rCH']:
        for harm in harmonics:
            d1 = next((d for d in all_data if d['cycle'] == cycle and d['harm'] == harm and d['bfc'] == 'BFC1'), None)
            d2 = next((d for d in all_data if d['cycle'] == cycle and d['harm'] == harm and d['bfc'] == 'BFC2'), None)
            
            if not d1 and not d2: continue
            
            chart = ScatterChart()
            chart.title = f"{cycle} - {harm}"
            chart.style = 13
            chart.x_axis.title = 'RPM'
            chart.y_axis.title = 'dB'
            chart.x_axis.scaling.min = 1000
            chart.x_axis.scaling.max = 4000
            chart.y_axis.scaling.min = 120
            chart.y_axis.scaling.max = 138
            
            idx1 = all_data.index(d1) + 2 if d1 else -1
            idx2 = all_data.index(d2) + 2 if d2 else -1
            
            xvalues = Reference(ws_data, min_col=4, min_row=1, max_col=4+len(target_rpms)-1)
            
            if idx1 != -1:
                values = Reference(ws_data, min_col=4, min_row=idx1, max_col=4+len(target_rpms)-1)
                series = Series(values, xvalues, title_from_data=False, title='BFC1')
                chart.series.append(series)
                
            if idx2 != -1:
                values = Reference(ws_data, min_col=4, min_row=idx2, max_col=4+len(target_rpms)-1)
                series = Series(values, xvalues, title_from_data=False, title='BFC2')
                chart.series.append(series)
                
            cell = ws_charts.cell(row=chart_row, column=chart_col)
            ws_charts.add_chart(chart, cell.coordinate)
            
            chart_col += 10
            if chart_col > 30:
                chart_col = 1
                chart_row += 15

    output_path = r"C:\Users\Louis\Downloads\Export_NVH_Courbes.xlsx"
    wb.save(output_path)
    print(f"Saved to {output_path}")

if __name__ == '__main__':
    main()
