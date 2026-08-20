import cv2
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

try:
    import easyocr
    READER = easyocr.Reader(['en'], gpu=False)
except:
    READER = None

def get_plot_area(image):
    # Fallback method if OCR fails
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return 0, 0, image.shape[1], image.shape[0]
    
    max_area = 0
    best_rect = (0, 0, image.shape[1], image.shape[0])
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        area = w * h
        if area > max_area and area > 0.1 * image.shape[0] * image.shape[1]:
            max_area = area
            best_rect = (x, y, w, h)
            
    if max_area == 0:
        h, w = image.shape[:2]
        return int(w*0.1), int(h*0.1), int(w*0.8), int(h*0.8)
    return best_rect

def calibrate_axes(img_path):
    img = cv2.imread(img_path)
    if img is None: return None
    
    h, w = img.shape[:2]
    # Enlarge the image for better OCR and finer analysis
    scale_factor = 3
    img_large = cv2.resize(img, (w * scale_factor, h * scale_factor), interpolation=cv2.INTER_CUBIC)
    
    if READER is None: return None
    
    import re
    results = READER.readtext(img_large)
    
    x_labels = []
    y_labels = []
    
    for (bbox, text, prob) in results:
        clean_text = re.sub(r'[^\d.]', '', text)
        if not clean_text: continue
        try:
            val = float(clean_text)
        except:
            continue
            
        cx = (bbox[0][0] + bbox[2][0]) / 2 / scale_factor
        cy = (bbox[0][1] + bbox[2][1]) / 2 / scale_factor
        
        if cx < w * 0.15 and 50 <= val <= 200:
            y_labels.append((cy, val))
            
        if cy > h * 0.8 and 500 <= val <= 10000:
            x_labels.append((cx, val))
            
    x_scale, y_scale = None, None
    x_offset, y_offset = None, None
    x_min_val, x_max_val = 1000, 4000
    y_min_val, y_max_val = 120, 138
    
    if len(x_labels) >= 2:
        x_labels.sort(key=lambda x: x[0])
        dx_px = x_labels[-1][0] - x_labels[0][0]
        dx_val = x_labels[-1][1] - x_labels[0][1]
        if dx_val != 0:
            x_scale = dx_val / dx_px
            x_offset = x_labels[0][0] - x_labels[0][1] / x_scale
            x_min_val = x_labels[0][1]
            x_max_val = x_labels[-1][1]
            
    if len(y_labels) >= 2:
        y_labels.sort(key=lambda x: x[1], reverse=True) # highest value should be smallest y
        y_labels.sort(key=lambda x: x[0])
        dy_px = y_labels[-1][0] - y_labels[0][0]
        dy_val = y_labels[-1][1] - y_labels[0][1]
        if dy_px != 0:
            y_scale = dy_val / dy_px
            y_offset = y_labels[0][0] - y_labels[0][1] / y_scale
            y_min_val = min([v[1] for v in y_labels])
            y_max_val = max([v[1] for v in y_labels])

    # If OCR failed to find enough labels, use fallback
    if x_scale is None or y_scale is None:
        x_plot, y_plot, w_plot, h_plot = get_plot_area(img)
        x_scale = (4000 - 1000) / w_plot
        x_offset = x_plot - 1000 / x_scale
        y_scale = (120 - 138) / h_plot # Note: 138 is top (smaller y)
        y_offset = y_plot - 138 / y_scale
        
    return {
        'x_scale': x_scale,
        'x_offset': x_offset,
        'y_scale': y_scale,
        'y_offset': y_offset,
        'x_min': x_min_val,
        'x_max': x_max_val
    }

def extract_red_curve_v2(img_path, calib):
    img = cv2.imread(img_path)
    if img is None: return [], []
    
    h, w = img.shape[:2]
    # Enlarge the image 4x for very fine pixel extraction
    scale_factor = 4
    img_large = cv2.resize(img, (w * scale_factor, h * scale_factor), interpolation=cv2.INTER_CUBIC)
    
    hsv = cv2.cvtColor(img_large, cv2.COLOR_BGR2HSV)
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    
    lower_red2 = np.array([170, 70, 50])
    upper_red2 = np.array([180, 255, 255])
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    
    mask = mask1 | mask2
    
    points = np.where(mask > 0)
    if len(points[0]) == 0:
        return [], []
        
    # convert back to original scale
    pts_x = points[1] / scale_factor
    pts_y = points[0] / scale_factor
    
    rpms = (pts_x - calib['x_offset']) * calib['x_scale']
    dbs = (pts_y - calib['y_offset']) * calib['y_scale']
    
    max_db_at_rpm = {}
    for r, d in zip(rpms, dbs):
        r_bin = round(r)
        if r_bin not in max_db_at_rpm or d > max_db_at_rpm[r_bin]:
            max_db_at_rpm[r_bin] = d
            
    if not max_db_at_rpm: return [], []
    
    r_keys = sorted(list(max_db_at_rpm.keys()))
    d_vals = [max_db_at_rpm[k] for k in r_keys]
    
    return r_keys, d_vals

def main():
    slides = {
        4: ('BFC1', 'rDL'),
        5: ('BFC1', 'rDH'),
        6: ('BFC1', 'rCL'),
        7: ('BFC1', 'rCH'),
        9: ('BFC2', 'rDL'),
        10: ('BFC2', 'rDH'),
        11: ('BFC2', 'rCL'),
        12: ('BFC2', 'rCH')
    }
    
    harmonics = ['H6', 'H7.4', 'H14.8', 'H18', 'H28', 'H32', 'H54', 'H56']
    
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Donnees_Brutes"
    
    # We will compute a dynamic set of RPMs based on max min
    all_target_rpms = set()
    
    all_data = []
    img_dir = "extracted_images"
    
    # Pre-calibrate per cycle (assume all 8 images in a cycle have same axes)
    cycle_calibs = {}
    
    for slide_num, (bfc, cycle) in slides.items():
        if cycle not in cycle_calibs:
            print(f"Calibrating for cycle {cycle}...")
            first_img = os.path.join(img_dir, f"slide_{slide_num}_img_1.png")
            if os.path.exists(first_img):
                calib = calibrate_axes(first_img)
                if calib:
                    cycle_calibs[cycle] = calib
                    print(f"  {cycle} calibration: X_min={calib['x_min']}, X_max={calib['x_max']}")
                else:
                    cycle_calibs[cycle] = None
                    
        for i, harm in enumerate(harmonics):
            img_name = f"slide_{slide_num}_img_{i+1}.png"
            img_path = os.path.join(img_dir, img_name)
            
            calib = cycle_calibs.get(cycle)
            
            if os.path.exists(img_path) and calib:
                # generate target rpms for this cycle
                x_min = max(0, int(calib['x_min']))
                x_max = int(calib['x_max'])
                # Start from the next multiple of 250
                start_rpm = x_min + (250 - x_min % 250) if x_min % 250 != 0 else x_min
                target_rpms = list(range(start_rpm, x_max + 250, 250))
                all_target_rpms.update(target_rpms)
                
                r_keys, d_vals = extract_red_curve_v2(img_path, calib)
                if r_keys:
                    interpolated_dbs = np.interp(target_rpms, r_keys, d_vals, left=np.nan, right=np.nan)
                    dbs = [round(val, 1) if not np.isnan(val) else None for val in interpolated_dbs]
                else:
                    dbs = [None]*len(target_rpms)
            else:
                target_rpms = []
                dbs = []
                
            all_data.append({
                'bfc': bfc,
                'cycle': cycle,
                'harm': harm,
                'target_rpms': target_rpms,
                'dbs': dbs
            })
            
    # Write headers
    sorted_rpms = sorted(list(all_target_rpms))
    headers = ['BFC', 'Cycle', 'Harmonique'] + [f"{rpm} RPM" for rpm in sorted_rpms]
    ws_data.append(headers)
    
    for row_data in all_data:
        row = [row_data['bfc'], row_data['cycle'], row_data['harm']]
        db_map = dict(zip(row_data['target_rpms'], row_data['dbs']))
        for rpm in sorted_rpms:
            row.append(db_map.get(rpm, None))
        ws_data.append(row)
            
    ws_charts = wb.create_sheet(title="Graphiques")
    
    chart_col = 1
    chart_row = 1
    
    for cycle in ['rDL', 'rDH', 'rCL', 'rCH']:
        for harm in harmonics:
            d1 = next((d for d in all_data if d['cycle'] == cycle and d['harm'] == harm and d['bfc'] == 'BFC1'), None)
            d2 = next((d for d in all_data if d['cycle'] == cycle and d['harm'] == harm and d['bfc'] == 'BFC2'), None)
            
            if not d1 and not d2: continue
            
            chart = ScatterChart()
            chart.title = f"{cycle} - {harm}"
            chart.style = 13
            chart.x_axis.title = 'RPM'
            chart.y_axis.title = 'dB'
            
            idx1 = all_data.index(d1) + 2 if d1 else -1
            idx2 = all_data.index(d2) + 2 if d2 else -1
            
            xvalues = Reference(ws_data, min_col=4, min_row=1, max_col=4+len(sorted_rpms)-1)
            
            if idx1 != -1 and any(v is not None for v in d1['dbs']):
                values = Reference(ws_data, min_col=4, min_row=idx1, max_col=4+len(sorted_rpms)-1)
                series = Series(values, xvalues, title_from_data=False)
                series.tx = Reference(ws_data, min_col=1, min_row=idx1) # reference BFC cell for title
                chart.series.append(series)
                
            if idx2 != -1 and any(v is not None for v in d2['dbs']):
                values = Reference(ws_data, min_col=4, min_row=idx2, max_col=4+len(sorted_rpms)-1)
                series = Series(values, xvalues, title_from_data=False)
                series.tx = Reference(ws_data, min_col=1, min_row=idx2)
                chart.series.append(series)
                
            cell = ws_charts.cell(row=chart_row, column=chart_col)
            ws_charts.add_chart(chart, cell.coordinate)
            
            chart_col += 10
            if chart_col > 30:
                chart_col = 1
                chart_row += 15

    output_path = r"C:\Users\Louis\Downloads\Export_NVH_Courbes_V2.xlsx"
    wb.save(output_path)
    print(f"Saved to {output_path}")

if __name__ == '__main__':
    main()
